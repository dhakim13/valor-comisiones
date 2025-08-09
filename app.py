import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime

st.set_page_config(page_title="Valor Telecom - Comisiones", page_icon="📊", layout="wide")
st.title("📊 Generador de Comisiones | Valor Telecom")
st.caption("Carga la base mensual y la plantilla del distribuidor. Calcula comisiones por ventana y esquema, y las vuelca sobre la plantilla.")

# =========================
# Helpers de columnas / limpieza
# =========================
def normalize_dn(series: pd.Series) -> pd.Series:
    out = series.astype(str).str.replace(r"\.0$", "", regex=True)
    def fix(x: str) -> str:
        try:
            s = str(x)
            if "e+" in s.lower():
                return str(int(float(s)))
            return s.split(".")[0]
        except Exception:
            return s
    return out.apply(fix)

def get_col(df: pd.DataFrame, candidates) -> str:
    for c in candidates:
        if c in df.columns:
            return c
    df_map = {str(c).strip().lower(): c for c in df.columns}
    for c in candidates:
        key = str(c).strip().lower()
        if key in df_map:
            return df_map[key]
    raise KeyError(f"No se encontró ninguna de estas columnas: {candidates}")

# =========================
# Reglas de negocio
# =========================
def cartera_pct_mbb(n_altas_mes: int) -> float:
    if n_altas_mes < 50:
        return 0.03
    elif n_altas_mes < 300:
        return 0.05
    elif n_altas_mes < 500:
        return 0.07
    elif n_altas_mes < 1000:
        return 0.08
    else:
        return 0.10

def days_in_month(ts: pd.Timestamp) -> int:
    # Días del mes natural del timestamp ts
    return int((ts + pd.offsets.MonthEnd(0)).day)

def ventana_por_dias(d: int) -> str:
    # d = días desde activación (1-based)
    if d <= 0:
        return "FUERA"
    if 1 <= d <= 30:
        return "M"
    if 31 <= d <= 60:
        return "M+1"
    if 61 <= d <= 90:
        return "M+2"
    if 91 <= d <= 365:
        return "M3–12"
    return "FUERA"

def clasificar_producto_basico(row: pd.Series) -> str:
    """
    Clasifica solo el producto base: MBB / MiFi / HBB
    """
    tipo = str(row.get("TIPO", "")).upper()
    costo = row.get("COSTO PAQUETE", np.nan)
    if "MOB" in tipo:
        return "MBB"
    if costo in [110, 120, 160, 245, 375, 480, 620]:
        return "MiFi"
    if costo in [99, 115, 349, 399, 439, 500]:
        return "HBB"
    # fallback
    return "MBB"

def clasificar_esquema(row: pd.Series) -> str:
    """
    Devuelve uno de:
    - "MBB"
    - "MiFi-Equipo"
    - "MiFi-SoloSIM"
    - "MiFi-10GB-1REC"
    - "HBB-Equipo"
    - "HBB-SoloSIM"

    Heurística por PLAN (ajústala cuando nos des las reglas exactas).
    """
    prod = row.get("PRODUCTO_BASE", "MBB")
    plan = str(row.get("PLAN", "")).upper()

    if prod == "MBB":
        return "MBB"

    if prod == "MiFi":
        if "10GB" in plan or "10 GB" in plan:
            return "MiFi-10GB-1REC"
        if "EQUIPO" in plan or "KIT" in plan:
            return "MiFi-Equipo"
        if "SIM" in plan:
            return "MiFi-SoloSIM"
        return "MiFi-SoloSIM"  # fallback

    if prod == "HBB":
        if "SIM" in plan:
            return "HBB-SoloSIM"
        if "EQUIPO" in plan or "FIJO" in plan or "ROUTER" in plan:
            return "HBB-Equipo"
        return "HBB-Equipo"  # fallback

    return "MBB"

def min_recharge_for_scheme(esquema: str) -> int:
    if esquema.startswith("MiFi"):
        # MiFi-10GB-1REC tiene mínimo distinto para cartera (120)
        return 120 if esquema == "MiFi-10GB-1REC" else 110
    if esquema.startswith("HBB"):
        return 99
    if esquema == "MBB":
        return 35
    return 0

def pct_cartera_for_scheme_window(esquema: str, ventana: str, pct_mbb_global: float) -> float:
    # MBB usa pct_mbb_global
    if esquema == "MBB":
        return pct_mbb_global
    # MiFi Equipo: 10%/15%/10%/5% (M / M+1 / M+2 / M3–12)
    if esquema == "MiFi-Equipo":
        return {"M": 0.10, "M+1": 0.15, "M+2": 0.10, "M3–12": 0.05}.get(ventana, 0.0)
    # MiFi Solo SIM y MiFi-10GB-1REC: 5% M1–12
    if esquema in ("MiFi-SoloSIM", "MiFi-10GB-1REC"):
        return 0.05 if ventana in ("M", "M+1", "M+2", "M3–12") else 0.0
    # HBB (Equipo / Solo SIM): 5% M1–12
    if esquema.startswith("HBB"):
        return 0.05 if ventana in ("M", "M+1", "M+2", "M3–12") else 0.0
    return 0.0

# =========================
# Motor de cálculo
# =========================
def calc_report(df_tot: pd.DataFrame, df_rec: pd.DataFrame, dist_name: str, year: int, month: int) -> BytesIO:
    month_start = pd.Timestamp(year, month, 1)
    month_end   = month_start + pd.offsets.MonthEnd(1)
    dias_mes    = days_in_month(month_start)

    # Columnas clave (tolerante a variantes)
    colF_TOT = get_col(df_tot, ["FECHA"])
    colDN_TOT = get_col(df_tot, ["DN"])
    colPLAN = get_col(df_tot, ["PLAN"])
    colCOSTO = get_col(df_tot, ["COSTO PAQUETE", "COSTO PAQ", "COSTO"])
    colTIPO = get_col(df_tot, ["TIPO"])
    colDIST = get_col(df_tot, ["DISTRIBUIDOR ", "DISTRIBUIDOR"])

    colF_REC = get_col(df_rec, ["FECHA"])
    colDN_REC = get_col(df_rec, ["DN"])
    colMONTO = get_col(df_rec, ["MONTO", "MONTO RECARGA", "RECARGA"])
    colFPAGO = get_col(df_rec, ["FORMA DE PAGO", "FP", "PAGO"])

    # Normalización
    tot = df_tot.copy()
    rec = df_rec.copy()
    tot[colF_TOT] = pd.to_datetime(tot[colF_TOT], errors="coerce")
    rec[colF_REC] = pd.to_datetime(rec[colF_REC], errors="coerce")
    tot["DN_NORM"] = normalize_dn(tot[colDN_TOT])
    rec["DN_NORM"] = normalize_dn(rec[colDN_REC])

    # Filtrar distribuidor
    mask_dist = tot[colDIST].astype(str).str.strip().str.lower() == dist_name.strip().lower()
    tot_d = tot[mask_dist].copy()
    dns_d = set(tot_d["DN_NORM"].dropna())

    # Activaciones del mes (para % MBB)
    altas_mes = tot_d[(tot_d[colF_TOT] >= month_start) & (tot_d[colF_TOT] <= month_end)].copy()
    n_altas_mes = int(altas_mes["DN_NORM"].nunique())
    pct_mbb = cartera_pct_mbb(n_altas_mes)

    # Producto/esquema por línea
    if "TIPO" not in tot_d.columns:
        tot_d["TIPO"] = tot_d[colTIPO]
    if "COSTO PAQUETE" not in tot_d.columns:
        tot_d["COSTO PAQUETE"] = tot_d[colCOSTO]
    tot_d["PRODUCTO_BASE"] = tot_d.apply(clasificar_producto_basico, axis=1)
    tot_d["ESQUEMA"] = tot_d.apply(clasificar_esquema, axis=1)

    # Base de líneas únicas del distribuidor
    base_lineas = tot_d[[colDN_TOT, "DN_NORM", colPLAN, colF_TOT, "PRODUCTO_BASE", "ESQUEMA", "COSTO PAQUETE"]].rename(
        columns={colDN_TOT: "DN", colF_TOT: "FECHA_ALTA", colPLAN: "PLAN"}
    )

    # Recargas del mes del universo del distribuidor
    rec_m = rec[(rec[colF_REC] >= month_start) & (rec[colF_REC] <= month_end)].copy()
    rec_m = rec_m[rec_m["DN_NORM"].isin(dns_d)].copy()

    # Join recargas con líneas (para conocer alta, esquema, etc.)
    det = rec_m.merge(base_lineas, on="DN_NORM", how="left")

    # Ventana por recarga (días desde alta)
    det["DIAS_DESDE_ALTA"] = (det[colF_REC] - det["FECHA_ALTA"]).dt.days + 1
    det["VENTANA"] = det["DIAS_DESDE_ALTA"].apply(lambda d: ventana_por_dias(int(d)) if pd.notnull(d) else "FUERA")

    # Mínimo por esquema (por recarga)
    det["MINIMO_ESQUEMA"] = det["ESQUEMA"].apply(min_recharge_for_scheme)

    # % cartera por esquema/ventana (MBB usa pct global)
    det["PCT_CARTERA"] = det.apply(
        lambda r: pct_cartera_for_scheme_window(str(r["ESQUEMA"]), str(r["VENTANA"]), pct_mbb),
        axis=1
    )

    # Elegibilidad por mínimo del esquema en esa recarga (nota: MBB >= 35)
    det["ELEGIBLE"] = det[colMONTO] >= det["MINIMO_ESQUEMA"]

    # Comisión por recarga (solo si elegible y dentro de ventanas válidas)
    det["COMISION_$"] = np.where(
        det["ELEGIBLE"] & det["PCT_CARTERA"].gt(0),
        det[colMONTO] * det["PCT_CARTERA"],
        0.0
    ).round(2)

    # Bono MiFi 10GB + 1ª recarga = $50 si 1ª recarga cae en día 31–60
    # Detectar 1ª recarga por línea (en toda la vida) y ver si su ventana es M+1
    # Para eso necesitamos el historial de recarga completo. Si solo tenemos el mes,
    # usamos heurística: 1ª recarga observada en el dataset y DIAS_DESDE_ALTA entre 31–60.
    # (Cuando nos des la hoja de detalle histórico global, lo mejoramos).
    det_sorted = rec[rec["DN_NORM"].isin(dns_d)].copy()
    det_sorted = det_sorted.merge(base_lineas[["DN_NORM", "FECHA_ALTA", "ESQUEMA"]], on="DN_NORM", how="left")
    det_sorted["DIAS_DESDE_ALTA"] = (det_sorted[colF_REC] - det_sorted["FECHA_ALTA"]).dt.days + 1
    det_sorted = det_sorted.sort_values([ "DN_NORM", colF_REC ])
    first_rec = det_sorted.groupby("DN_NORM", as_index=False).first()[["DN_NORM", colF_REC, "DIAS_DESDE_ALTA", "ESQUEMA"]]
    first_rec["BONO_50"] = (first_rec["ESQUEMA"] == "MiFi-10GB-1REC") & first_rec["DIAS_DESDE_ALTA"].between(31, 60, inclusive="both")
    # Unimos a detalle del MES y pagamos $50 SOLO si la 1ª recarga ocurrió dentro del mes y cumple
    det = det.merge(first_rec[["DN_NORM", colF_REC, "BONO_50"]].rename(columns={colF_REC: "FECHA_PRIMERA_REC"}), on="DN_NORM", how="left")
    det["BONO_50_PAGADO"] = np.where(
        det["BONO_50"].fillna(False) & det[colF_REC].eq(det["FECHA_PRIMERA_REC"]) &
        (det[colF_REC] >= month_start) & (det[colF_REC] <= month_end),
        50.0, 0.0
    )
    det["COMISION_TOTAL_RECARGA_$"] = (det["COMISION_$"] + det["BONO_50_PAGADO"]).round(2)

    # ======= Agregados por línea (para ANEXO) =======
    rec_suma_por_linea = (
        det.groupby(["DN_NORM"], as_index=False)
        .agg(
            RECARGA_TOTAL_MES=(colMONTO, "sum"),
            COMISION_TOTAL_MES=("COMISION_TOTAL_RECARGA_$", "sum")
        )
    )

    anexo = base_lineas.merge(rec_suma_por_linea, on="DN_NORM", how="left")
    anexo["RECARGA_TOTAL_MES"] = anexo["RECARGA_TOTAL_MES"].fillna(0.0)
    anexo["COMISION_TOTAL_MES"] = anexo["COMISION_TOTAL_MES"].fillna(0.0).round(2)

    # ======= RESUMEN (1 fila) =======
    # Elegibilidad MBB para bono por volumen: recarga >= 35 y activo todo el mes natural
    # (activo todo el mes ⇒ fecha de alta <= inicio del mes)
    tot_mbb = base_lineas[base_lineas["ESQUEMA"] == "MBB"].copy()
    if not tot_mbb.empty:
        # Recarga por línea en mes
        rec_mbb = det[det["ESQUEMA"] == "MBB"].groupby("DN_NORM", as_index=False)[colMONTO].sum().rename(columns={colMONTO: "REC_MES"})
        tot_mbb = tot_mbb.merge(rec_mbb, on="DN_NORM", how="left").fillna({"REC_MES": 0.0})
        tot_mbb["ACTIVO_TODO_MES"] = tot_mbb["FECHA_ALTA"] <= month_start
        tot_mbb["ELEGIBLE_BONO"] = (tot_mbb["REC_MES"] >= 35) & (tot_mbb["ACTIVO_TODO_MES"])
        lineas_mbb_elegibles = int(tot_mbb["ELEGIBLE_BONO"].sum())
    else:
        lineas_mbb_elegibles = 0

    resumen = pd.DataFrame([{
        "Distribuidor": dist_name,
        "Mes": f'{month_start.strftime("%B").capitalize()} {year}',
        "Altas del mes": n_altas_mes,
        "Líneas MBB elegibles (mes)": lineas_mbb_elegibles,
        "Recargas totales del mes ($)": round(float(det[colMONTO].sum() if not det.empty else 0.0), 2),
        "Porcentaje Cartera aplicado (MBB)": float(pct_mbb),
        "Comisión total del mes ($)": round(float(anexo["COMISION_TOTAL_MES"].sum()), 2)
    }])

    # ======= RESUMEN MES por esquema =======
   resumen_mes = (
    anexo.groupby("ESQUEMA", as_index=False)
    .agg({
        "DN_NORM": "nunique",
        "RECARGA_TOTAL_MES": "sum",
        "COMISION_TOTAL_MES": "sum",
    })
    .rename(columns={
        "DN_NORM": "Lineas",
        "RECARGA_TOTAL_MES": "Recarga_Mes_$",
        "COMISION_TOTAL_MES": "Comision_Mes_$",
    })
)

    # ======= HISTORIAL DE ACTIVACIONES (solo mes) =======
    hist = altas_mes[[colF_TOT, "DN_NORM", colPLAN, colCOSTO, "ESQUEMA"]].rename(
        columns={colF_TOT: "FECHA", "DN_NORM": "DN", colCOSTO: "COSTO PAQUETE", colPLAN: "PLAN"}
    ).sort_values("FECHA")

    # ======= CARTERA MES (detalle por recarga con comisión) =======
    cartera_mes = det[[colF_REC, "DN_NORM", colPLAN, colMONTO, colFPAGO, "ESQUEMA", "VENTANA", "ELEGIBLE", "PCT_CARTERA", "BONO_50_PAGADO", "COMISION_TOTAL_RECARGA_$"]].rename(
        columns={colF_REC: "FECHA", "DN_NORM": "DN", colMONTO: "MONTO", colFPAGO: "FORMA DE PAGO", colPLAN: "PLAN"}
    ).sort_values(["FECHA", "DN"])

    # Export preliminar en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as w:
        resumen.to_excel(w, sheet_name="RESUMEN", index=False)
        anexo.to_excel(w, sheet_name="ANEXO", index=False)
        hist.to_excel(w, sheet_name="HISTORIAL DE ACTIVACIONES", index=False)
        resumen_mes.to_excel(w, sheet_name=f'RESUMEN {month_start.strftime("%B").upper()} {year}', index=False)
        cartera_mes.to_excel(w, sheet_name=f'CARTERA {month_start.strftime("%B").upper()} {year}', index=False)
    output.seek(0)
    return output

# =========================
# Volcado sobre plantilla
# =========================
def write_like_template(template_bytes, new_sheets: dict) -> BytesIO:
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows

    if isinstance(template_bytes, BytesIO):
        template_bytes.seek(0)
        wb = openpyxl.load_workbook(template_bytes)
    else:
        wb = openpyxl.load_workbook(template_bytes)

    for sh_name, df in new_sheets.items():
        if sh_name in wb.sheetnames:
            ws_old = wb[sh_name]
            wb.remove(ws_old)
        ws = wb.create_sheet(sh_name)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        # ancho básico
        for col_cells in ws.columns:
            try:
                col_letter = col_cells[0].column_letter
                max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
                ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)
            except Exception:
                pass

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# =========================
# UI
# =========================
col1, col2 = st.columns(2)
with col1:
    base_file = st.file_uploader("1) Base mensual (VT Reporte Comercial...)", type=["xlsx"])
    plantilla_file = st.file_uploader("2) Plantilla/Historial del distribuidor (.xlsx)", type=["xlsx"])
    st.caption("Base: 'Desgloce Totales' (header fila 2) y 'Desgloce Recarga' (header fila 4). La plantilla se preserva salvo las 5 hojas del mes.")
with col2:
    st.write("Parámetros")
    dist = st.text_input("Distribuidor", value="ActivateCel")
    year = st.number_input("Año", min_value=2023, max_value=2100, value=2025, step=1)
    month = st.number_input("Mes (1–12)", min_value=1, max_value=12, value=6, step=1)

st.divider()
st.markdown("**Paso a paso:** 1) Sube ambos archivos. 2) Verifica distribuidor/año/mes. 3) Clic en **Generar reporte**. 4) Descarga el Excel final.")

if base_file is not None and plantilla_file is not None and st.button("Generar reporte"):
    try:
        xls = pd.ExcelFile(base_file, engine="openpyxl")
        if "Desgloce Totales" not in xls.sheet_names or "Desgloce Recarga" not in xls.sheet_names:
            st.error("El archivo base debe contener las hojas 'Desgloce Totales' y 'Desgloce Recarga'.")
        else:
            df_tot = pd.read_excel(base_file, sheet_name="Desgloce Totales", header=1, engine="openpyxl")
            df_rec = pd.read_excel(base_file, sheet_name="Desgloce Recarga", header=3, engine="openpyxl")

            buf_calc = calc_report(df_tot, df_rec, dist.strip(), int(year), int(month))
            month_str = datetime(int(year), int(month), 1).strftime("%B").upper()

            with pd.ExcelFile(buf_calc, engine="openpyxl") as xf:
                df_resumen  = pd.read_excel(xf, sheet_name="RESUMEN")
                df_anexo    = pd.read_excel(xf, sheet_name="ANEXO")
                df_hist     = pd.read_excel(xf, sheet_name="HISTORIAL DE ACTIVACIONES")
                df_res_mes  = pd.read_excel(xf, sheet_name=f"RESUMEN {month_str} {int(year)}")
                df_cart_mes = pd.read_excel(xf, sheet_name=f"CARTERA {month_str} {int(year)}")

            new_sheets = {
                "RESUMEN": df_resumen,
                "ANEXO": df_anexo,
                "HISTORIAL DE ACTIVACIONES": df_hist,
                f"RESUMEN {month_str} {int(year)}": df_res_mes,
                f"CARTERA {month_str} {int(year)}": df_cart_mes
            }

            final_buf = write_like_template(plantilla_file, new_sheets)
            fname = f"COMISION VALOR DISTRIBUIDOR {dist.upper()} {month_str} {int(year)}.xlsx"

            st.success("✅ Reporte armado con comisiones por esquema y ventana.")
            st.download_button("⬇️ Descargar Excel final", data=final_buf, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        st.exception(e)
